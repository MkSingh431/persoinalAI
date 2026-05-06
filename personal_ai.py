#!/usr/bin/env python3
"""Privacy-first personal desktop assistant.

This tool is intentionally conservative: it can search, summarize, draft,
open apps, and organize files, but file-moving operations require explicit
confirmation and it never sends messages.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Iterable, Sequence


APP_DIR = Path(__file__).resolve().parent
LOG_DIR = APP_DIR / "assistant_logs"
REMINDERS_FILE = APP_DIR / "reminders.json"
CONFIG_FILE = APP_DIR / "assistant_config.json"

TEXT_EXTENSIONS = {
    ".bat",
    ".cmd",
    ".csv",
    ".ini",
    ".json",
    ".log",
    ".md",
    ".ps1",
    ".py",
    ".rtf",
    ".text",
    ".toml",
    ".tsv",
    ".txt",
    ".xml",
    ".yaml",
    ".yml",
}

DOCUMENT_EXTENSIONS = {
    ".doc",
    ".docx",
    ".pdf",
    ".ppt",
    ".pptx",
    ".xls",
    ".xlsx",
} | TEXT_EXTENSIONS

SENSITIVE_WORDS = {
    "aadhaar",
    "account",
    "bank",
    "credential",
    "cvv",
    "otp",
    "pan",
    "passport",
    "password",
    "secret",
    "tax",
    "upi",
}


def now_iso() -> str:
    return dt.datetime.now().astimezone().isoformat(timespec="seconds")


def log_action(action: str, details: dict) -> None:
    LOG_DIR.mkdir(exist_ok=True)
    record = {"time": now_iso(), "action": action, "details": details}
    with (LOG_DIR / "actions.jsonl").open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=True) + "\n")


def print_table(rows: Sequence[Sequence[str]], headers: Sequence[str]) -> None:
    if not rows:
        print("No results.")
        return
    widths = [len(header) for header in headers]
    for row in rows:
        for idx, value in enumerate(row):
            widths[idx] = min(max(widths[idx], len(str(value))), 90)

    def crop(value: str, width: int) -> str:
        value = str(value)
        return value if len(value) <= width else value[: width - 3] + "..."

    header_line = "  ".join(crop(header, widths[idx]).ljust(widths[idx]) for idx, header in enumerate(headers))
    print(header_line)
    print("  ".join("-" * width for width in widths))
    for row in rows:
        print("  ".join(crop(str(value), widths[idx]).ljust(widths[idx]) for idx, value in enumerate(row)))


def resolve_root(raw_root: str | None) -> Path:
    root = Path(raw_root or ".").expanduser().resolve()
    if not root.exists():
        raise SystemExit(f"Folder does not exist: {root}")
    if not root.is_dir():
        raise SystemExit(f"Root is not a folder: {root}")
    return root


def iter_files(root: Path, include_hidden: bool = False) -> Iterable[Path]:
    ignored_names = {
        "$recycle.bin",
        ".git",
        ".hg",
        ".svn",
        "__pycache__",
        "node_modules",
        "venv",
        ".venv",
    }
    stack = [root]
    while stack:
        current = stack.pop()
        try:
            with os.scandir(current) as entries:
                for entry in entries:
                    name_lower = entry.name.lower()
                    if not include_hidden and entry.name.startswith("."):
                        continue
                    if entry.is_dir(follow_symlinks=False):
                        if name_lower not in ignored_names:
                            stack.append(Path(entry.path))
                    elif entry.is_file(follow_symlinks=False):
                        yield Path(entry.path)
        except (PermissionError, FileNotFoundError, OSError):
            continue


def is_sensitive(path: Path) -> bool:
    haystack = str(path).lower()
    return any(word in haystack for word in SENSITIVE_WORDS)


def human_size(size: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{size} B"


def format_mtime(path: Path) -> str:
    try:
        return dt.datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    except OSError:
        return "unknown"


def command_search(args: argparse.Namespace) -> None:
    root = resolve_root(args.root)
    query = args.query.lower()
    allowed_exts = {ext.lower() if ext.startswith(".") else "." + ext.lower() for ext in args.ext}
    results: list[tuple[Path, str]] = []

    for path in iter_files(root, include_hidden=args.hidden):
        if allowed_exts and path.suffix.lower() not in allowed_exts:
            continue
        reason = ""
        if query in path.name.lower():
            reason = "name"
        elif args.content and path.suffix.lower() in TEXT_EXTENSIONS:
            try:
                text = path.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                text = ""
            if query in text.lower():
                reason = "content"
        if reason:
            results.append((path, reason))
            if len(results) >= args.limit:
                break

    table_rows = []
    for path, reason in results:
        try:
            size = human_size(path.stat().st_size)
        except OSError:
            size = "unknown"
        table_rows.append((str(path), reason, format_mtime(path), size, "yes" if is_sensitive(path) else ""))
    print_table(table_rows, ("Path", "Match", "Modified", "Size", "Sensitive"))
    log_action("search", {"root": str(root), "query": args.query, "results": len(results)})


def command_recent(args: argparse.Namespace) -> None:
    root = resolve_root(args.root)
    cutoff = dt.datetime.now().timestamp() - (args.days * 24 * 60 * 60)
    files: list[Path] = []
    for path in iter_files(root, include_hidden=args.hidden):
        try:
            if path.stat().st_mtime >= cutoff:
                files.append(path)
        except OSError:
            continue
    files.sort(key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True)
    rows = []
    for path in files[: args.limit]:
        rows.append((str(path), format_mtime(path), human_size(path.stat().st_size), "yes" if is_sensitive(path) else ""))
    print_table(rows, ("Path", "Modified", "Size", "Sensitive"))
    log_action("recent", {"root": str(root), "days": args.days, "results": min(len(files), args.limit)})


def read_text_file(path: Path, max_chars: int) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")[:max_chars]


def read_pdf(path: Path, max_chars: int) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError:
        return "PDF support requires pypdf. Install with: pip install -r requirements.txt"
    try:
        reader = PdfReader(str(path))
        chunks = []
        for page in reader.pages:
            chunks.append(page.extract_text() or "")
            if sum(len(chunk) for chunk in chunks) >= max_chars:
                break
        return "\n".join(chunks)[:max_chars]
    except Exception as exc:  # noqa: BLE001 - show friendly extraction failure.
        return f"Could not read PDF: {exc}"


def read_docx(path: Path, max_chars: int) -> str:
    try:
        import docx  # type: ignore
    except ImportError:
        return "Word support requires python-docx. Install with: pip install -r requirements.txt"
    try:
        document = docx.Document(str(path))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)[:max_chars]
    except Exception as exc:  # noqa: BLE001
        return f"Could not read Word document: {exc}"


def read_xlsx(path: Path, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return "Excel support requires openpyxl. Install with: pip install -r requirements.txt"
    try:
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        lines = []
        for sheet in workbook.worksheets:
            lines.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(max_row=20, values_only=True):
                values = ["" if value is None else str(value) for value in row]
                lines.append(" | ".join(values))
                if sum(len(line) for line in lines) >= max_chars:
                    return "\n".join(lines)[:max_chars]
        return "\n".join(lines)[:max_chars]
    except Exception as exc:  # noqa: BLE001
        return f"Could not read Excel workbook: {exc}"


def extract_text(path: Path, max_chars: int) -> str:
    suffix = path.suffix.lower()
    if suffix in TEXT_EXTENSIONS:
        return read_text_file(path, max_chars)
    if suffix == ".pdf":
        return read_pdf(path, max_chars)
    if suffix == ".docx":
        return read_docx(path, max_chars)
    if suffix == ".xlsx":
        return read_xlsx(path, max_chars)
    return f"Unsupported file type for summary: {suffix or 'no extension'}"


def simple_summary(text: str, max_sentences: int = 5) -> str:
    clean = re.sub(r"\s+", " ", text).strip()
    if not clean:
        return "No readable text found."
    if clean.startswith(("PDF support requires", "Word support requires", "Excel support requires", "Could not read", "Unsupported")):
        return clean
    sentences = re.split(r"(?<=[.!?])\s+", clean)
    selected = sentences[:max_sentences]
    summary = " ".join(sentence.strip() for sentence in selected if sentence.strip())
    if not summary:
        summary = clean[:900]
    keywords = top_keywords(clean)
    if keywords:
        summary += "\n\nKeywords: " + ", ".join(keywords[:10])
    return summary


def top_keywords(text: str) -> list[str]:
    stop_words = {
        "about",
        "after",
        "also",
        "and",
        "are",
        "but",
        "for",
        "from",
        "has",
        "have",
        "into",
        "not",
        "that",
        "the",
        "this",
        "with",
        "you",
        "your",
    }
    words = re.findall(r"[A-Za-z][A-Za-z0-9_-]{2,}", text.lower())
    counts: dict[str, int] = {}
    for word in words:
        if word in stop_words:
            continue
        counts[word] = counts.get(word, 0) + 1
    return [word for word, _count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:10]]


def command_summarize(args: argparse.Namespace) -> None:
    path = Path(args.path).expanduser().resolve()
    if not path.exists() or not path.is_file():
        raise SystemExit(f"File does not exist: {path}")
    if is_sensitive(path) and not args.yes:
        print(f"Sensitive-looking path detected: {path}")
        answer = input("Type SUMMARIZE to continue: ").strip()
        if answer != "SUMMARIZE":
            print("Cancelled.")
            return
    text = extract_text(path, args.max_chars)
    print(simple_summary(text, args.sentences))
    log_action("summarize", {"path": str(path), "sensitive": is_sensitive(path)})


def open_target(target: str) -> None:
    target_lower = target.lower().strip()
    aliases = {
        "chrome": "chrome",
        "edge": "msedge",
        "excel": "excel",
        "notepad": "notepad",
        "powerpoint": "powerpnt",
        "whatsapp": "https://web.whatsapp.com",
        "whatsapp-web": "https://web.whatsapp.com",
        "word": "winword",
    }
    resolved = aliases.get(target_lower, target)
    if sys.platform.startswith("win"):
        subprocess.Popen(["cmd", "/c", "start", "", resolved], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", resolved], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    else:
        subprocess.Popen(["xdg-open", resolved], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def command_open(args: argparse.Namespace) -> None:
    opened = []
    for target in args.targets:
        open_target(target)
        opened.append(target)
    print("Opened: " + ", ".join(opened))
    if any(target.lower() in {"whatsapp", "whatsapp-web"} for target in opened):
        print("WhatsApp Web opened. I will only draft replies; sending requires your approval.")
    log_action("open", {"targets": opened})


def command_draft_reply(args: argparse.Namespace) -> None:
    message = args.message.strip()
    context = args.context.strip() if args.context else ""
    tone = args.tone
    greeting = "Hello"
    if tone == "friendly":
        opener = f"{greeting}, thanks for your message."
    elif tone == "brief":
        opener = "Noted, thank you."
    else:
        opener = f"{greeting}, thank you for reaching out."

    lower_message = message.lower()
    if any(word in lower_message for word in ["meeting", "call", "schedule", "appointment"]):
        body = "Please share the preferred time and agenda, and I will confirm availability."
    elif any(word in lower_message for word in ["urgent", "asap", "immediately", "today"]):
        body = "I will review this on priority and get back to you shortly."
    elif any(word in lower_message for word in ["invoice", "payment", "quote", "quotation"]):
        body = "I will check the details and respond with the required information."
    elif context:
        body = f"Regarding {context}, I will review the details and respond shortly."
    else:
        body = "I will check the details and respond shortly."

    closing = "Regards" if tone == "professional" else "Thanks"
    draft = f"{opener} {body}\n\n{closing}"
    print(draft)
    print("\nDraft only. Nothing was sent.")
    log_action("draft_reply", {"tone": tone, "message_chars": len(message), "context": context})


def load_reminders() -> list[dict]:
    if not REMINDERS_FILE.exists():
        return []
    try:
        return json.loads(REMINDERS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []


def save_reminders(reminders: list[dict]) -> None:
    REMINDERS_FILE.write_text(json.dumps(reminders, indent=2, ensure_ascii=True), encoding="utf-8")


def parse_when(raw_when: str) -> str:
    raw_when = raw_when.strip()
    shortcuts = {
        "today": dt.datetime.now().replace(hour=18, minute=0, second=0, microsecond=0),
        "tomorrow": (dt.datetime.now() + dt.timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0),
    }
    if raw_when.lower() in shortcuts:
        return shortcuts[raw_when.lower()].isoformat(timespec="minutes")
    formats = ["%Y-%m-%d %H:%M", "%Y-%m-%d", "%d-%m-%Y %H:%M", "%d-%m-%Y"]
    for fmt in formats:
        try:
            parsed = dt.datetime.strptime(raw_when, fmt)
            if "%H:%M" not in fmt:
                parsed = parsed.replace(hour=9, minute=0)
            return parsed.isoformat(timespec="minutes")
        except ValueError:
            continue
    raise SystemExit("Use when as 'YYYY-MM-DD HH:MM', 'YYYY-MM-DD', 'today', or 'tomorrow'.")


def command_remind(args: argparse.Namespace) -> None:
    reminders = load_reminders()
    next_id = max((item.get("id", 0) for item in reminders), default=0) + 1
    reminder = {
        "id": next_id,
        "text": args.text,
        "when": parse_when(args.when),
        "status": "open",
        "created": now_iso(),
    }
    reminders.append(reminder)
    save_reminders(reminders)
    print(f"Reminder added #{next_id}: {reminder['when']} - {args.text}")
    log_action("reminder_add", {"id": next_id, "when": reminder["when"]})


def command_reminders(args: argparse.Namespace) -> None:
    reminders = load_reminders()
    if args.open_only:
        reminders = [item for item in reminders if item.get("status") == "open"]
    rows = [(str(item.get("id")), item.get("when", ""), item.get("status", ""), item.get("text", "")) for item in reminders]
    print_table(rows, ("ID", "When", "Status", "Text"))
    log_action("reminders_list", {"count": len(reminders), "open_only": args.open_only})


def command_complete_reminder(args: argparse.Namespace) -> None:
    reminders = load_reminders()
    changed = False
    for item in reminders:
        if item.get("id") == args.id:
            item["status"] = "completed"
            item["completed"] = now_iso()
            changed = True
            break
    if not changed:
        raise SystemExit(f"Reminder not found: {args.id}")
    save_reminders(reminders)
    print(f"Reminder #{args.id} marked completed.")
    log_action("reminder_complete", {"id": args.id})


def month_folder_for(path: Path) -> str:
    modified = dt.datetime.fromtimestamp(path.stat().st_mtime)
    return modified.strftime("%Y-%m")


def looks_like_screenshot(path: Path) -> bool:
    if path.suffix.lower() not in {".png", ".jpg", ".jpeg", ".webp", ".bmp"}:
        return False
    name = path.stem.lower()
    return "screenshot" in name or "screen shot" in name or name.startswith("snip")


def command_organize_screenshots(args: argparse.Namespace) -> None:
    root = resolve_root(args.root)
    destination = Path(args.destination).expanduser().resolve() if args.destination else root / "Screenshots"
    moves: list[tuple[Path, Path]] = []
    planned_targets: set[Path] = set()
    for path in iter_files(root, include_hidden=False):
        if destination in path.parents:
            continue
        if looks_like_screenshot(path):
            target_dir = destination / month_folder_for(path)
            target = target_dir / path.name
            suffix_counter = 1
            while target.exists() or target in planned_targets:
                target = target_dir / f"{path.stem}-{suffix_counter}{path.suffix}"
                suffix_counter += 1
            planned_targets.add(target)
            moves.append((path, target))

    rows = [(str(source), str(target)) for source, target in moves[: args.limit]]
    print_table(rows, ("Source", "Target"))
    if len(moves) > args.limit:
        print(f"...and {len(moves) - args.limit} more.")

    if not args.execute:
        print("\nDry run only. Add --execute to move files.")
        log_action("organize_screenshots_dry_run", {"root": str(root), "matches": len(moves)})
        return

    if not moves:
        print("No screenshots to move.")
        return

    print(f"\nThis will move {len(moves)} file(s).")
    answer = input("Type ORGANIZE to continue: ").strip()
    if answer != "ORGANIZE":
        print("Cancelled.")
        log_action("organize_screenshots_cancelled", {"root": str(root), "matches": len(moves)})
        return

    for source, target in moves:
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(source), str(target))
    print(f"Moved {len(moves)} screenshot file(s) into {destination}.")
    log_action("organize_screenshots_execute", {"root": str(root), "destination": str(destination), "moved": len(moves)})


def parse_since(raw_since: str) -> dt.datetime:
    value = raw_since.lower().strip()
    today = dt.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if value == "today":
        return today
    if value == "yesterday":
        return today - dt.timedelta(days=1)
    try:
        return dt.datetime.strptime(raw_since, "%Y-%m-%d")
    except ValueError as exc:
        raise SystemExit("Use since as 'today', 'yesterday', or 'YYYY-MM-DD'.") from exc


def command_daily_report(args: argparse.Namespace) -> None:
    root = resolve_root(args.root)
    since = parse_since(args.since)
    changed: list[Path] = []
    for path in iter_files(root, include_hidden=False):
        try:
            if dt.datetime.fromtimestamp(path.stat().st_mtime) >= since:
                changed.append(path)
        except OSError:
            continue
    changed.sort(key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True)

    open_reminders = [item for item in load_reminders() if item.get("status") == "open"]
    lines = [
        f"# Daily Work Report - {dt.datetime.now().strftime('%Y-%m-%d')}",
        "",
        f"Workspace: {root}",
        f"Since: {since.strftime('%Y-%m-%d %H:%M')}",
        "",
        "## Recently Modified Files",
    ]
    if changed:
        for path in changed[: args.limit]:
            lines.append(f"- {format_mtime(path)} | {human_size(path.stat().st_size)} | {path}")
    else:
        lines.append("- No modified files found.")
    lines.extend(["", "## Open Reminders"])
    if open_reminders:
        for item in open_reminders:
            lines.append(f"- #{item.get('id')} | {item.get('when')} | {item.get('text')}")
    else:
        lines.append("- No open reminders.")

    report = "\n".join(lines)
    print(report)
    if args.save:
        reports_dir = APP_DIR / "reports"
        reports_dir.mkdir(exist_ok=True)
        output = reports_dir / f"daily_report_{dt.datetime.now().strftime('%Y-%m-%d')}.md"
        output.write_text(report, encoding="utf-8")
        print(f"\nSaved report: {output}")
    log_action("daily_report", {"root": str(root), "since": args.since, "files": min(len(changed), args.limit), "saved": args.save})


def command_logs(args: argparse.Namespace) -> None:
    log_file = LOG_DIR / "actions.jsonl"
    if not log_file.exists():
        print("No logs yet.")
        return
    lines = log_file.read_text(encoding="utf-8").splitlines()[-args.limit :]
    rows = []
    for line in lines:
        try:
            item = json.loads(line)
        except json.JSONDecodeError:
            continue
        rows.append((item.get("time", ""), item.get("action", ""), json.dumps(item.get("details", {}), ensure_ascii=True)))
    print_table(rows, ("Time", "Action", "Details"))


def write_default_config() -> None:
    if CONFIG_FILE.exists():
        return
    config = {
        "assistant_name": "Personal AI",
        "privacy_rules": [
            "Never send messages automatically.",
            "Ask before deleting, moving, or modifying sensitive files.",
            "Keep action logs in assistant_logs/actions.jsonl.",
        ],
        "sensitive_keywords": sorted(SENSITIVE_WORDS),
    }
    CONFIG_FILE.write_text(json.dumps(config, indent=2, ensure_ascii=True), encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Personal AI desktop assistant for local productivity workflows.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    search = subparsers.add_parser("search", help="Search files by name, optionally text content.")
    search.add_argument("query", help="Text to search for.")
    search.add_argument("--root", default=".", help="Folder to search.")
    search.add_argument("--limit", type=int, default=50, help="Maximum results.")
    search.add_argument("--ext", nargs="*", default=[], help="Optional file extensions such as .pdf .docx .xlsx.")
    search.add_argument("--content", action="store_true", help="Also search inside text files.")
    search.add_argument("--hidden", action="store_true", help="Include hidden folders.")
    search.set_defaults(func=command_search)

    recent = subparsers.add_parser("recent", help="Show recently modified files.")
    recent.add_argument("--root", default=".", help="Folder to inspect.")
    recent.add_argument("--days", type=int, default=7, help="Lookback window in days.")
    recent.add_argument("--limit", type=int, default=50, help="Maximum results.")
    recent.add_argument("--hidden", action="store_true", help="Include hidden folders.")
    recent.set_defaults(func=command_recent)

    summarize = subparsers.add_parser("summarize", help="Summarize a local document.")
    summarize.add_argument("path", help="File path.")
    summarize.add_argument("--max-chars", type=int, default=20000, help="Maximum extracted characters.")
    summarize.add_argument("--sentences", type=int, default=5, help="Summary sentence count.")
    summarize.add_argument("--yes", action="store_true", help="Skip sensitive-path confirmation.")
    summarize.set_defaults(func=command_summarize)

    open_parser = subparsers.add_parser("open", help="Open apps, folders, files, or URLs.")
    open_parser.add_argument("targets", nargs="+", help="Examples: chrome whatsapp excel notepad C:\\Users\\...")
    open_parser.set_defaults(func=command_open)

    draft = subparsers.add_parser("draft-reply", help="Draft a WhatsApp/email reply without sending.")
    draft.add_argument("--message", required=True, help="Incoming message text.")
    draft.add_argument("--context", default="", help="Optional context for the reply.")
    draft.add_argument("--tone", choices=["professional", "friendly", "brief"], default="professional", help="Reply style.")
    draft.set_defaults(func=command_draft_reply)

    remind = subparsers.add_parser("remind", help="Add a reminder.")
    remind.add_argument("text", help="Reminder text.")
    remind.add_argument("--when", required=True, help="YYYY-MM-DD HH:MM, YYYY-MM-DD, today, or tomorrow.")
    remind.set_defaults(func=command_remind)

    reminders = subparsers.add_parser("reminders", help="List reminders.")
    reminders.add_argument("--open-only", action="store_true", help="Show only open reminders.")
    reminders.set_defaults(func=command_reminders)

    complete = subparsers.add_parser("complete-reminder", help="Mark a reminder completed.")
    complete.add_argument("id", type=int, help="Reminder ID.")
    complete.set_defaults(func=command_complete_reminder)

    organize = subparsers.add_parser("organize-screenshots", help="Group screenshots by month.")
    organize.add_argument("--root", default=".", help="Folder to scan.")
    organize.add_argument("--destination", default="", help="Destination folder. Defaults to <root>/Screenshots.")
    organize.add_argument("--limit", type=int, default=100, help="Preview rows.")
    organize.add_argument("--execute", action="store_true", help="Actually move files after confirmation.")
    organize.set_defaults(func=command_organize_screenshots)

    report = subparsers.add_parser("daily-report", help="Create a simple work report from local files and reminders.")
    report.add_argument("--root", default=".", help="Folder to inspect.")
    report.add_argument("--since", default="today", help="today, yesterday, or YYYY-MM-DD.")
    report.add_argument("--limit", type=int, default=40, help="Maximum modified files in report.")
    report.add_argument("--save", action="store_true", help="Save report under reports/.")
    report.set_defaults(func=command_daily_report)

    logs = subparsers.add_parser("logs", help="Show completed action history.")
    logs.add_argument("--limit", type=int, default=20, help="Number of log entries.")
    logs.set_defaults(func=command_logs)

    return parser


def main() -> None:
    write_default_config()
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
